# -------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      СашаНастя
#
# Created:     24.10.2019
# Copyright:   (c) СашаНастя 2019
# Licence:     <your licence>
# -------------------------------------------------------------------------------
from openpyxl.worksheet.worksheet import Worksheet

from xlFile import xlLabel
import os
from openpyxl import load_workbook, Workbook

def mains():
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = 'Fuck'


    ws.cell(1, 1).value = 10

    wb.save("my.xlsx")


if __name__ == '__main__':
    mains()
