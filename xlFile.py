from PyQt5.QtCore import QDir
from openpyxl.styles import Font, colors, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins, PrintOptions, PrintPageSetup
from openpyxl import Workbook


class xlLabel(object):

    def __init__(self):
        self.colsVal = ['A', 'B', 'C', 'D', 'E']
        self.rowVal = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
        self.fontType = Font(name='Calibri', sz=11, color=colors.BLACK, bold=True)
        self.filename = 'наклейки.xlsx'

        self.wb = Workbook()
        self.sheet = self.wb.active
        self.sheet.title = 'Наклейки'
        self.rutilink = "\nРуТиЛинк\n +7 (961) 518-33-65"
        self.margins_tblr = 0.1968  # 0.393701/2
        self.value = [59.60, 18.90]

    def getValue(self, name=''):

        _dict_ = {
            1: "RTL00000",
            2: "RTL0000",
            3: "RTL000",
            4: "RTL00",
            5: "RTL0",
            6: "RTL",
        }

        temp = int(name)

        if len(name) in _dict_:
            for i in range(1, 14):
                for j in range(1, 6):
                    self.sheet.cell(i, j).value = _dict_[len(name)] + str(temp) + self.rutilink
                    temp = temp + 1

        word_wrap_string = Alignment(wrapText=True,
                                     horizontal="center",
                                     vertical='center')

        double_border_side = Side(border_style='dotted')

        square_border = Border(top=double_border_side,
                               right=double_border_side,
                               bottom=double_border_side,
                               left=double_border_side)

        self.sheet.page_margins = PageMargins(left=self.margins_tblr,
                                              right=self.margins_tblr,
                                              top=self.margins_tblr,
                                              bottom=self.margins_tblr)

        self.sheet.sheet_properties.pageSetUpPr.fitToPage = True
        self.sheet.print_area = "A1:E13"
        self.sheet.page_setup = PrintPageSetup(worksheet=self.sheet,
                                               orientation='portrait',
                                               paperSize=self.sheet.PAPERSIZE_A4,
                                               fitToHeight=1,
                                               fitToWidth=1,
                                               scale=100,
                                               horizontalDpi=300,
                                               verticalDpi=300)

        self.sheet.print_options = PrintOptions(horizontalCentered=True,
                                                verticalCentered=True)

        for i in range(1, 6):
            for j in range(1, 14):
                self.sheet.cell(j, i).border = square_border
                self.sheet.cell(j, i).font = self.fontType
                self.sheet.cell(j, i).alignment = word_wrap_string

        for cols in self.colsVal:
            for row in self.rowVal:
                self.sheet.column_dimensions[cols].width = self.value[1]
                self.sheet.row_dimensions[row].height = self.value[0]

        self.wb.save(QDir.homePath() + '/Desktop/' + self.filename)
        self.wb.close()
